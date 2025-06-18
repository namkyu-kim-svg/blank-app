import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from datetime import datetime
import os
from employee_manager import employee_manager

def parse_range(range_str):
    """셀 범위 문자열을 파싱하여 시작과 끝 좌표를 반환"""
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        return min_col, min_row, max_col, max_row
    except:
        return None, None, None, None

def calculate_auto_dimensions(total_rows=50, total_cols=12, margin_inches=0.65):
    """A4 용지 크기에 맞춰 자동으로 행 높이와 열 너비 계산"""
    
    # A4 용지 크기 (인치)
    A4_WIDTH_INCH = 8.27
    A4_HEIGHT_INCH = 11.69
    
    # 사용 가능한 영역 계산 (여백 제외)
    usable_width = A4_WIDTH_INCH - (margin_inches * 2)
    usable_height = A4_HEIGHT_INCH - (margin_inches * 2)
    
    # 페이지당 행 수
    rows_per_page = 25
    
    # 행 높이 계산 (포인트 단위)
    row_height_points = (usable_height * 72) / 25
    
    # 열 너비 계산 (엑셀 단위)
    col_width_excel_units = (usable_width * 7.5) / total_cols
    
    # 최소/최대 제한
    row_height_points = max(35, min(50, row_height_points))
    col_width_excel_units = max(8, min(18, col_width_excel_units))
    
    return {
        'row_height': round(row_height_points, 1),
        'col_width': round(col_width_excel_units, 1),
        'usable_width_inch': round(usable_width, 2),
        'usable_height_inch': round(usable_height, 2),
        'rows_per_page': rows_per_page
    }

def create_advanced_business_trip_report(employees_data, additional_costs, filename="출장복명서.xlsx"):
    """
    출장복명서 생성 - 여러 직원의 자동 계산 기능
    
    Args:
        employees_data: 직원별 출장 데이터 리스트
        additional_costs: 추가 비용 데이터 리스트
        filename: 출력 파일명
    
    Returns:
        str: 생성된 파일의 전체 경로
    """
    
    # 새 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "출장복명서"
    
    # 기본 헤더 정보 설정
    current_date = datetime.now().strftime('%Y년 %m월 %d일')
    
    # 기본 셀 데이터 매핑 (13행까지는 기존과 동일)
    basic_cell_data = {
        'A1:L1': '출 장 복 명 서',
        'A4:B4': '소       속',
        'C4:F4': '㈜엔이비',
        'G2:H2': '담당자',
        'I2:J2': '책임연구원',
        'K2:L2': '대표이사',
        'G3:H3': '',  # 빈 병합 셀 (서명용)
        'I3:J3': '',  # 빈 병합 셀 (서명용)
        'K3:L3': '',  # 빈 병합 셀 (서명용)
        'G4:H4': '과제책임자',
        'I4:L4': employees_data[0]['project_manager'] if employees_data else '',
        'A5:B5': '연구과제명',
        'C5:L5': employees_data[0]['project_name'] if employees_data else '',
        'A6:B6': '출발시간',
        'A7:B7': '도착시간',
        'G6:H7': '출장지',
        'I6:L7': employees_data[0]['destination'] if employees_data else '',
        'A8:B8': '출장결과',
        'C8:L8': employees_data[0]['trip_purpose'] if employees_data else '',
        'A9:L9': '(단위:원)',
        'A10:L10': '지 급 신 청',
        'A11:A12': '성명',
        'B11:B12': '결제구분',
        'D11:E12': '직책',
        'F11:H11': '일비',
        'F12': '일',
        'G12': '일당',
        'H12': '금액',
        'I11:K11': '식비',
        'I12': '일',
        'J12': '일당',
        'K12': '금액',
        'L11:L12': '소계',
    }
    
    # 기본 헤더 정보 입력
    for range_str, value in basic_cell_data.items():
        if not value and range_str not in ['G3:H3', 'I3:J3', 'K3:L3']:
            continue
            
        min_col, min_row, max_col, max_row = parse_range(range_str)
        if min_col is None:
            continue
            
        # 셀 병합 처리
        if min_col != max_col or min_row != max_row:
            try:
                ws.merge_cells(start_row=min_row, start_column=min_col,
                             end_row=max_row, end_column=max_col)
            except Exception as e:
                print(f"셀 병합 실패 {range_str}: {e}")
        
        # 값 입력 (서명용 셀 제외)
        if range_str not in ['G3:H3', 'I3:J3', 'K3:L3']:
            cell = ws.cell(row=min_row, column=min_col)
            cell.value = value
    
    # 추가 비용 데이터 처리 - 같은 항목끼리 합산 (13행부터 시작)
    current_row = 13
    total_additional_cost = 0
    
    # 비용 항목별로 그룹화하여 합산 (수식으로 저장)
    cost_groups = {}
    for cost_data in additional_costs:
        item_name = cost_data['item']
        if item_name not in cost_groups:
            cost_groups[item_name] = {
                'amounts': [],  # 개별 금액들을 리스트로 저장
                'payment_methods': []
            }
        cost_groups[item_name]['amounts'].append(cost_data['amount'])
        # 결제 방식 중복 제거하여 저장
        if cost_data['payment_method'] not in cost_groups[item_name]['payment_methods']:
            cost_groups[item_name]['payment_methods'].append(cost_data['payment_method'])
    
    # 그룹화된 비용 항목들을 엑셀에 입력
    for item_name, group_data in cost_groups.items():
        # A열: 비용 항목
        ws.cell(row=current_row, column=1, value=item_name)
        
        # B열: 결제 방식 (여러 개인 경우 줄바꿈으로 구분)
        payment_methods_text = '\n'.join(group_data['payment_methods'])
        ws.cell(row=current_row, column=2, value=payment_methods_text)
        
        # L열: 수식으로 합산된 금액
        amounts = group_data['amounts']
        if len(amounts) == 1:
            # 단일 항목인 경우 그냥 값으로 입력
            ws.cell(row=current_row, column=12, value=amounts[0])
            total_additional_cost += amounts[0]
        else:
            # 여러 항목인 경우 수식으로 입력
            formula = "=" + "+".join(str(amount) for amount in amounts)
            ws.cell(row=current_row, column=12, value=formula)
            total_additional_cost += sum(amounts)
        
        current_row += 1
    
    # 동적 직원 데이터 입력 (current_row부터 시작)
    total_employee_cost = 0
    
    for emp_data in employees_data:
        # A열: 성명
        ws.cell(row=current_row, column=1, value=emp_data['employee_name'])
        
        # B열: 결제구분
        ws.cell(row=current_row, column=2, value="계좌이체")
        
        # D-E열: 직책 (병합)
        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
        ws.cell(row=current_row, column=4, value=emp_data['position'])
        
        # F열: 일비 일수
        ws.cell(row=current_row, column=6, value=emp_data['trip_days'])
        
        # G열: 일비 일당
        ws.cell(row=current_row, column=7, value=emp_data['daily_allowance_per_day'])
        
        # H열: 일비 금액 (출장일수 * 일비) - 수식으로 입력
        daily_formula = f"=F{current_row}*G{current_row}"
        ws.cell(row=current_row, column=8, value=daily_formula)
        
        # I열: 식비 일수
        ws.cell(row=current_row, column=9, value=emp_data['trip_days'])
        
        # J열: 식비 일당
        ws.cell(row=current_row, column=10, value=emp_data['meal_cost_per_day'])
        
        # K열: 식비 금액 (출장일수 * 식비) - 수식으로 입력
        meal_formula = f"=I{current_row}*J{current_row}"
        ws.cell(row=current_row, column=11, value=meal_formula)
        
        # L열: 소계 (H + K) - 수식으로 입력
        total_formula = f"=H{current_row}+K{current_row}"
        ws.cell(row=current_row, column=12, value=total_formula)
        
        # 계산용으로 row_total 유지
        row_total = emp_data['daily_allowance_total'] + emp_data['meal_cost_total']
        total_employee_cost += row_total
        current_row += 1
    
    total_cost_sum = total_employee_cost + total_additional_cost
    
    # 출장 날짜 정보 추가 (첫 번째 직원 기준)
    if employees_data:
        first_emp = employees_data[0]
        start_datetime = f"{first_emp['start_date'].strftime('%Y년  %m월 %d일')} {first_emp['start_time'].strftime('%H시 %M분')}"
        end_datetime = f"{first_emp['end_date'].strftime('%Y년  %m월 %d일')} {first_emp['end_time'].strftime('%H시 %M분')}"
        
        ws.cell(row=6, column=3, value=start_datetime)  # C6
        ws.cell(row=7, column=3, value=end_datetime)    # C7
        ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=6)  # C6:F6
        ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=6)  # C7:F7
    
    # 합계 행 추가 - 수식으로 처리
    total_row = current_row + 1
    ws.cell(row=total_row, column=1, value='합계')
    
    # 합계 수식 생성 (13행부터 current_row까지의 L열 합계)
    if current_row > 13:
        total_formula = f"=SUM(L13:L{current_row})"
        ws.cell(row=total_row, column=12, value=total_formula)
    else:
        ws.cell(row=total_row, column=12, value=0)
    
    # 하단 정보 추가 (빈칸 줄 없음)
    bottom_row = current_row + 2
    ws.merge_cells(start_row=bottom_row, start_column=1, end_row=bottom_row, end_column=12)
    ws.cell(row=bottom_row, column=1, value='위와 같이 출장 복명서를 제출합니다.')
    
    ws.merge_cells(start_row=bottom_row+1, start_column=1, end_row=bottom_row+1, end_column=12)
    ws.cell(row=bottom_row+1, column=1, value=current_date)
    
    ws.merge_cells(start_row=bottom_row+2, start_column=1, end_row=bottom_row+2, end_column=12)
    ws.cell(row=bottom_row+2, column=1, value='주 식 회 사 엔 이 비')
    
    # 추가수당 섹션 추가
    extra_row = bottom_row + 4
    ws.merge_cells(start_row=extra_row, start_column=1, end_row=extra_row+2, end_column=2)
    ws.cell(row=extra_row, column=1, value='추가수당\n(해당시)')
    
    # 추가수당 데이터 계산 (일수) - 첫 번째 직원 기준으로만
    total_holiday_days = employees_data[0].get('holiday_days', 0) if employees_data else 0
    total_special_days = employees_data[0].get('special_days', 0) if employees_data else 0
    total_dangerous_days = employees_data[0].get('dangerous_days', 0) if employees_data else 0
    
    # 휴일출장
    ws.merge_cells(start_row=extra_row, start_column=3, end_row=extra_row, end_column=6)
    ws.cell(row=extra_row, column=3, value='휴일출장 (토,일요일,공휴일)')
    ws.merge_cells(start_row=extra_row, start_column=7, end_row=extra_row, end_column=8)
    ws.cell(row=extra_row, column=7, value=f"{total_holiday_days}일" if total_holiday_days > 0 else "0")
    
    # 특수출장
    ws.merge_cells(start_row=extra_row+1, start_column=3, end_row=extra_row+1, end_column=6)
    ws.cell(row=extra_row+1, column=3, value='특수(선박탑승,해양조사)')
    ws.merge_cells(start_row=extra_row+1, start_column=7, end_row=extra_row+1, end_column=8)
    ws.cell(row=extra_row+1, column=7, value=f"{total_special_days}일" if total_special_days > 0 else "0")
    
    # 위험출장
    ws.merge_cells(start_row=extra_row+2, start_column=3, end_row=extra_row+2, end_column=6)
    ws.cell(row=extra_row+2, column=3, value='위험(다이빙,해저조사)')
    ws.merge_cells(start_row=extra_row+2, start_column=7, end_row=extra_row+2, end_column=8)
    ws.cell(row=extra_row+2, column=7, value=f"{total_dangerous_days}일" if total_dangerous_days > 0 else "0")
    
    # 추가수당 안내문
    ws.merge_cells(start_row=extra_row, start_column=10, end_row=extra_row+2, end_column=12)
    ws.cell(row=extra_row, column=10, value='*특수, 위험 출장시\n사진파일 첨부.\n*추가수당은 지급신청서에 기입하지 않음.')
    
    # 자동 크기 계산
    auto_dimensions = calculate_auto_dimensions(total_rows=50, total_cols=12, margin_inches=0.65)
    
    # 추가수당 섹션 아래 페이지 나누기 행 계산
    bottom_row = current_row + 2  # "위와 같이..." 시작 행
    extra_row = bottom_row + 4     # 추가수당 시작 행
    page_break_row = extra_row + 3 # 추가수당 섹션 끝 (3행에 걸쳐 있음)
    
    # 선박 승선 증빙 사진 첨부 안내 (특수/위험 출장시에만) - 2페이지 맨 처음에 배치
    if total_special_days > 0 or total_dangerous_days > 0:
        # 2페이지 시작점 = 페이지 나누기 행 + 1
        second_page_start = page_break_row + 1
        ws.merge_cells(start_row=second_page_start, start_column=1, end_row=second_page_start, end_column=12)
        ws.cell(row=second_page_start, column=1, value='* 2페이지에 선박 승선 증빙 사진 첨부')
    
    # 스타일 적용
    apply_advanced_styles(ws, auto_dimensions, current_row)
    
    # 페이지 설정 적용 (동적 페이지 나누기 포함)
    setup_page_settings_advanced(ws, page_break_row)
    
    # 출력 파일 경로
    output_path = os.path.join(os.getcwd(), filename)
    
    # 엑셀 파일 저장
    wb.save(output_path)
    
    return output_path

def create_business_trip_report(data, filename="출장보고서.xlsx"):
    """
    출장보고서 데이터를 받아서 엑셀 파일을 생성하는 함수
    
    Args:
        data: 출장보고서 데이터 딕셔너리
        filename: 출력 파일명
    
    Returns:
        str: 생성된 파일의 전체 경로
    """
    
    # 새 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "출장복명서"
    
    # 출장 날짜 포맷팅
    start_datetime = f"{data['start_date'].strftime('%Y년  %m월 %d일')} {data['start_time'].strftime('%H시 %M분')}"
    end_datetime = f"{data['end_date'].strftime('%Y년  %m월 %d일')} {data['end_time'].strftime('%H시 %M분')}"
    
    # 현재 날짜
    current_date = datetime.now().strftime('%Y년 %m월 %d일')
    
    # 셀 데이터 매핑
    cell_data = {
        'A1:L1': '출 장 복 명 서',
        'A4:B4': '소       속',
        'C4:F4': '㈜엔이비',
        'G2:H2': '담당자',
        'I2:J2': '책임연구원',
        'K2:L2': '대표이사',
        'G3:H3': '',  # 빈 병합 셀 (서명용)
        'I3:J3': '',  # 빈 병합 셀 (서명용)
        'K3:L3': '',  # 빈 병합 셀 (서명용)
        'G4:H4': '과제책임자',
        'I4:L4': data['project_manager'],
        'A5:B5': '연구과제명',
        'C5:L5': data['project_name'],
        'A6': '출발시간',
        'A7': '도착시간',
        'B6': '출발',
        'B7': '도착',
        'C6:F6': start_datetime,
        'C7:F7': end_datetime,
        'G6:H7': '출장지',
        'I6:L7': data['destination'],
        'A8:B8': '출장결과',
        'C8:L8': data['trip_purpose'],
        'A9:L9': '(단위:원)',
        'A10:L10': '지 급 신 청',
        'A11:A12': '성명',
        'B11:B12': '결재구분',
        'D11:F11': '일비',
        'G11:H11': '숙박료',
        'I11:K11': '식비',
        'L11:L12': '소계',

        'D13:F13': str(data['daily_allowance']) if data['daily_allowance'] > 0 else '',
        'G13:H13': str(data['accommodation']) if data['accommodation'] > 0 else '',
        'I13:K13': str(data['meal_cost']) if data['meal_cost'] > 0 else '',
        'L13:L13': str(data['daily_allowance'] + data['accommodation'] + data['meal_cost']),
        'A19:L19': '위와 같이 출장 복명서를 제출합니다.',
        'A20:L20': current_date,
        'A21:L21': '주 식 회 사 엔 이 비',
        'A22:B24': '추가수당\n(해당시)',
        'C22:F22': '휴일출장 (토,일요일,공휴일)',
        'G22:H22': str(data['holiday_amount']) if data['holiday_work'] else '0',
        'C23:F23': '특수(선박탑승,해양조사)',
        'G23:H23': str(data['special_amount']) if data['special_work'] else '0',
        'C24:F24': '위험(다이빙,해저조사)',
        'G24:H24': str(data['dangerous_amount']) if data['dangerous_work'] else '0',
        'J22:L24': '*특수, 위험 출장시\n사진파일 첨부.\n*추가수당은 지급신청서에 기입하지 않음.',
        'A26:L26': '* 선박 승선 증빙 사진 첨부' if (data['special_work'] or data['dangerous_work']) else ''
    }
    
    # 서명용 빈 셀들 (병합만 하고 값은 입력하지 않음)
    signature_cells = ['G3:H3', 'I3:J3', 'K3:L3']
    
    # 셀에 데이터 입력 및 병합
    for range_str, value in cell_data.items():
        min_col, min_row, max_col, max_row = parse_range(range_str)
        
        if min_col is None:
            continue
        
        # 서명용 셀들은 병합만 처리
        if range_str in signature_cells:
            # 셀 병합 처리
            if min_col != max_col or min_row != max_row:
                try:
                    ws.merge_cells(start_row=min_row, start_column=min_col,
                                 end_row=max_row, end_column=max_col)
                except Exception as e:
                    print(f"셀 병합 실패 {range_str}: {e}")
            continue  # 값은 입력하지 않고 다음으로
            
        # 일반 셀들은 빈 값이면 건너뛰기 (단, '0'은 제외)
        if not value and value != '0':
            continue
            
        # 먼저 셀 병합 처리
        if min_col != max_col or min_row != max_row:
            try:
                ws.merge_cells(start_row=min_row, start_column=min_col,
                             end_row=max_row, end_column=max_col)
            except Exception as e:
                print(f"셀 병합 실패 {range_str}: {e}")
        
        # 값 입력
        cell = ws.cell(row=min_row, column=min_col)
        cell.value = value
    
    # 자동 크기 계산
    auto_dimensions = calculate_auto_dimensions(total_rows=50, total_cols=12, margin_inches=0.65)
    
    # 스타일 적용
    apply_styles(ws, auto_dimensions)
    
    # 페이지 설정 적용
    setup_page_settings(ws)
    
    # 출력 파일 경로
    output_path = os.path.join(os.getcwd(), filename)
    
    # 엑셀 파일 저장
    wb.save(output_path)
    
    return output_path

def apply_styles(ws, auto_dimensions):
    """워크시트에 자동 계산된 크기로 스타일 적용"""
    
    # 기본 폰트 설정
    default_font = Font(name='맑은 고딕', size=10)
    
    # 제목 스타일
    title_font = Font(name='맑은 고딕', size=14, bold=True)
    
    # 테두리 스타일 (검은색)
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 표 영역에 모든 셀에 테두리 적용 (A1부터 L50까지)
    for row in range(1, 51):  # 1행부터 50행까지
        for col in range(1, 13):  # A열부터 L열까지
            cell = ws.cell(row=row, column=col)
            
            # A1:L1, A2:F2, A3:F3는 테두리 없음
            if (row == 1) or (row == 2 and col <= 6) or (row == 3 and col <= 6):  # A1:L1, A2:F2, A3:F3
                cell.border = None
            # 25행에는 굵은 테두리 적용 (페이지 구분선)
            elif row == 25:
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thick', color='000000')  # 하단만 굵게
                )
            else:
                cell.border = black_border
                
            # A9:L9 (단위:원)는 오른쪽 정렬
            if row == 9:
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
            cell.font = default_font
    
    # 제목 셀 스타일 적용
    if ws['A1'].value:
        ws['A1'].font = title_font
    
    # 자동 계산된 행 높이 적용
    for row_num in range(1, 51):
        if row_num == 9:  # A9:L9 (단위:원) 행은 높이를 작게
            ws.row_dimensions[row_num].height = auto_dimensions['row_height'] * 0.6  # 60%로 축소
        else:
            ws.row_dimensions[row_num].height = auto_dimensions['row_height']
    
    # 자동 계산된 열 너비 적용
    for col in range(1, 13):  # A~L열
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = auto_dimensions['col_width']

def setup_page_settings(ws):
    """페이지 설정 및 인쇄 영역 설정"""
    
    # 페이지 방향 설정 (세로)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    
    # 용지 크기 설정 (A4)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # 여백 설정 (단위: 인치)
    ws.page_margins = PageMargins(
        left=0.65, right=0.65, top=0.65, bottom=0.65,
        header=0.3, footer=0.3
    )
    
    # 페이지에 맞춤 설정
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 2  # 정확히 2페이지
    ws.page_setup.fitToWidth = 1   # 가로 1페이지
    
    # 인쇄영역 설정
    ws.print_area = 'A1:L50'
    
    # 페이지 나누기 (25행 이후)
    ws.row_breaks.append(Break(id=25))
    
    # 격자선 인쇄 설정
    ws.print_options.gridLines = False
    ws.print_options.gridLinesSet = True
    
    # 중앙 정렬
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    
    # 페이지 나누기 미리보기 설정
    ws.sheet_view.view = 'pageBreakPreview'
    ws.sheet_view.zoomScale = 100

def setup_page_settings_advanced(ws, page_break_row):
    """페이지 설정 및 인쇄 영역 설정 (동적 페이지 나누기)"""
    
    # 페이지 방향 설정 (세로)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    
    # 용지 크기 설정 (A4)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # 여백 설정 (단위: 인치)
    ws.page_margins = PageMargins(
        left=0.65, right=0.65, top=0.65, bottom=0.65,
        header=0.3, footer=0.3
    )
    
    # 페이지에 맞춤 설정
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 2  # 정확히 2페이지
    ws.page_setup.fitToWidth = 1   # 가로 1페이지
    
    # 인쇄영역 설정
    ws.print_area = 'A1:L50'
    
    # 동적 페이지 나누기 (추가수당 섹션 아래)
    ws.row_breaks.append(Break(id=page_break_row))
    
    # 격자선 인쇄 설정
    ws.print_options.gridLines = False
    ws.print_options.gridLinesSet = True
    
    # 중앙 정렬
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    
    # 페이지 나누기 미리보기 설정
    ws.sheet_view.view = 'pageBreakPreview'
    ws.sheet_view.zoomScale = 100

def apply_advanced_styles(ws, auto_dimensions, data_end_row):
    """스타일 적용"""
    
    # 다양한 폰트 설정
    default_font = Font(name='맑은 고딕', size=10)
    title_font = Font(name='맑은 고딕', size=24, bold=True)  # 1행 제목
    unit_font = Font(name='맑은 고딕', size=8)  # 9행 (단위:원)

    bottom_bold_font = Font(name='맑은 고딕', size=16, bold=True)  # 하단 텍스트들
    
    # 테두리 스타일 (검은색)
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 추가수당 섹션 아래 굵은 밑줄을 위한 행 계산
    bottom_row = data_end_row + 2  # "위와 같이..." 시작 행
    extra_row = bottom_row + 4     # 추가수당 시작 행
    page_break_row = extra_row + 3 # 추가수당 섹션 끝 (3행에 걸쳐 있음)
    
    # 표 영역에 모든 셀에 테두리 적용 (A1부터 L50까지)
    for row in range(1, 51):  # 1행부터 50행까지
        for col in range(1, 13):  # A열부터 L열까지
            cell = ws.cell(row=row, column=col)
            
            # A1:L1, A2:F2, A3:F3는 테두리 없음
            if (row == 1) or (row == 2 and col <= 6) or (row == 3 and col <= 6):
                cell.border = None
            # 데이터 영역 (13행부터 data_end_row까지)는 모든 테두리 적용
            elif 13 <= row <= data_end_row:
                cell.border = black_border
            # 추가수당 섹션 아래에 굵은 테두리 적용 (동적 페이지 구분선)
            elif row == page_break_row:
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thick', color='000000')  # 하단만 굵게
                )
            # 밑줄 아래(2페이지)는 테두리 없음
            elif row > page_break_row:
                cell.border = None
            else:
                cell.border = black_border
                
            # 정렬 설정
            if row == 9:  # A9:L9 (단위:원)는 오른쪽 정렬
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                cell.font = unit_font
            elif 13 <= row <= data_end_row:  # 데이터 행들은 중앙 정렬
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = default_font
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = default_font
    
    # 특별한 셀들의 폰트 설정
    if ws['A1'].value:  # 1행 제목
        ws['A1'].font = title_font
    
    # 6-7행 출발/도착 시간은 기본 폰트 사용 (별도 설정 불필요)
    
    # 하단 텍스트들 찾아서 bold 22pt 적용 (6-7행 제외)
    for row in range(1, 51):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                # 6-7행은 제외하고 적용
                if row not in [6, 7]:
                    if ('위와 같이 출장 복명서를 제출합니다' in cell.value or 
                        '년' in cell.value and '월' in cell.value and '일' in cell.value or
                        '주 식 회 사 엔 이 비' in cell.value):
                        cell.font = bottom_bold_font
    
    # 행별 높이 세밀 조정
    row_heights = {
        1: 40,    # 1행: 출장복명서 - 크게
        2: 18,    # 2행: 간격 줄임
        3: 50,    # 3행: 간격 늘림
        4: 18,    # 4행: 기본
        5: 25,    # 5행: 기본
        6: 18,    # 6행: 간격 많이 줄임
        7: 18,    # 7행: 간격 많이 줄임
        8: 60,    # 8행: 출장결과 - 4-5줄 여유있게
        9: 15,    # 9행: (단위:원) - 많이 줄임
        10: 20,   # 10행: 지급신청
        11: 18,   # 11행: 헤더 - 절반으로 줄임
        12: 18,   # 12행: 헤더 - 절반으로 줄임
    }
    
    # 13행부터 데이터 끝까지 18pt로 설정
    for row_num in range(13, data_end_row + 1):
        row_heights[row_num] = 20  # 데이터 행들 모두 18pt
    
    # 합계 행도 18pt
    row_heights[data_end_row + 1] = 20
    
    # 하단 문구들 간격 조정
    bottom_start_row = data_end_row + 2  # "위와 같이 출장 복명서를 제출합니다" 시작 행
    row_heights[bottom_start_row] = 22      # "위와 같이 출장 복명서를 제출합니다" 
    row_heights[bottom_start_row + 1] = 22  # 날짜
    row_heights[bottom_start_row + 2] = 22  # "주식회사 엔이비"
    
    # 행 높이 적용
    for row_num in range(1, 51):
        if row_num in row_heights:
            ws.row_dimensions[row_num].height = row_heights[row_num]
        else:
            ws.row_dimensions[row_num].height = auto_dimensions['row_height']
    
    # 자동 계산된 열 너비 적용
    for col in range(1, 13):  # A~L열
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = auto_dimensions['col_width']

def create_business_trip_application(application_data, filename="출장신청서.xlsx"):
    """
    출장신청서 생성 함수
    
    Args:
        application_data: 출장신청서 데이터 딕셔너리
        filename: 출력 파일명
    
    Returns:
        str: 생성된 파일의 전체 경로
    """
    
    # 새 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "출장신청서"
    
    # 현재 날짜
    current_date = datetime.now().strftime('%Y년 %m월 %d일')
    
    # 기본 셀 데이터 매핑
    cell_data = {
        'A1:G1': '출 장 신 청 서',
        'D2:D3': '결재',  # D2:D3 병합
        'E2': '담당', 
        'F2': '책임연구원',
        'G2': '대표이사',
        'E3': '',  # 서명용 빈 셀  
        'F3': '',  # 서명용 빈 셀
        'G3': '',  # 서명용 빈 셀
        'A4': '소속',
        'B4:D4': '㈜엔이비',
        'E4': '과제책임자',
        'F4:G4': application_data.get('project_manager', ''),
        'A5': '연구과제명',
        'B5:G5': application_data.get('project_name', ''),
        'A6': '출장기간',
        'B6:D6': application_data.get('trip_period', ''),
        'E6': '출장지',
        'F6:G6': application_data.get('destination', ''),
        'A7': '출장목적',
        'B7:G7': application_data.get('trip_purpose', ''),
        'A8:A9': '출장교통비',
        'B8:B9': '이용차량',
        'C8': '법인(차종)',
        'D8': application_data.get('company_car', ''),
        'E8:E9': '대중교통\n(항공,철도,\n선박)',
        'F8:G9': application_data.get('public_transport', ''),
        'C9': '자차(유종)',
        'A10:G10': '출 장 자',
        'A11:C11': '직급',
        'D11': '성명',
        'E11:F11': '계좌번호',
        'G11': '비고',
        'A20:G20': '위와 같이 출장을 신청합니다.',
        'A21:G21': current_date,
        'A22:G22': '주 식 회 사 엔 이 비',
    }
    
    # 기본 정보 입력
    for range_str, value in cell_data.items():
        min_col, min_row, max_col, max_row = parse_range(range_str)
        if min_col is None:
            continue
            
        # 셀 병합 처리
        if min_col != max_col or min_row != max_row:
            try:
                ws.merge_cells(start_row=min_row, start_column=min_col,
                             end_row=max_row, end_column=max_col)
            except Exception as e:
                print(f"셀 병합 실패 {range_str}: {e}")
        
        # 값 입력 (서명용 빈 셀 제외)
        if value:
            cell = ws.cell(row=min_row, column=min_col)
            cell.value = value
    
    # 출장자 데이터 입력 (12행부터 19행까지 - 최대 8명)
    travelers = application_data.get('travelers', [])
    
    # 8행 모두 병합 처리 (12-19행)
    for row_idx in range(8):  # 0~7 (8행)
        current_row = 12 + row_idx
        
        # A열: 직급 (A:C 병합)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        
        # E-F열: 계좌번호 (병합)
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        
        # 데이터가 있는 경우에만 값 입력
        if row_idx < len(travelers):
            traveler = travelers[row_idx]
            ws.cell(row=current_row, column=1, value=traveler.get('position', ''))
            ws.cell(row=current_row, column=4, value=traveler.get('name', ''))
            ws.cell(row=current_row, column=5, value=traveler.get('account', ''))
            ws.cell(row=current_row, column=7, value=traveler.get('note', ''))
    
    # 스타일 적용
    apply_application_styles(ws)
    
    # 페이지 설정 적용
    setup_application_page_settings(ws)
    
    # 출력 파일 경로
    output_path = os.path.join(os.getcwd(), filename)
    
    # 엑셀 파일 저장
    wb.save(output_path)
    
    return output_path

def apply_application_styles(ws):
    """출장신청서용 스타일 적용"""
    
    # 다양한 폰트 설정
    default_font = Font(name='맑은 고딕', size=11)
    title_font = Font(name='맑은 고딕', size=20, bold=True)  # 제목 크게
    header_font = Font(name='맑은 고딕', size=10, bold=True)  # 헤더
    section_header_font = Font(name='맑은 고딕', size=13, bold=True)  # 섹션 헤더 (출장자용)
    small_font = Font(name='맑은 고딕', size=9)  # 작은 글씨
    bottom_font = Font(name='맑은 고딕', size=14, bold=True)  # 하단 문구
    
    # 테두리 스타일
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 굵은 테두리 (결재란용)
    thick_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    # 모든 셀에 기본 스타일 적용 (A1부터 G22까지)
    for row in range(1, 23):
        for col in range(1, 8):  # A~G열
            cell = ws.cell(row=row, column=col)
            
            # A1 제목
            if row == 1:
                cell.border = None
                cell.font = title_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 2-3행 결재란 (테두리 bold 제거)
            elif 2 <= row <= 3 and 4 <= col <= 7:  # D2:G3
                cell.border = black_border  # thick_border → black_border로 변경
                cell.font = default_font  # bold 제거
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 4-11행 기본 정보 영역
            elif 4 <= row <= 11:
                cell.border = black_border
                # 라벨 셀들 (A열, E열의 특정 셀들)
                if (col == 1 and row in [4, 5, 6, 7, 11]) or (col == 5 and row in [4, 6]):
                    cell.font = header_font
                elif col == 1 and row == 10:  # "출장자" 섹션은 더 큰 폰트
                    cell.font = section_header_font
                else:
                    cell.font = default_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 12-19행 출장자 데이터 영역 (8명까지)
            elif 12 <= row <= 19:
                cell.border = black_border
                cell.font = default_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 20-22행 하단 영역 (출장자 8행으로 확장)
            elif 20 <= row <= 22:
                cell.border = None
                cell.font = bottom_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.border = None
                cell.font = default_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 특별한 셀들 폰트 설정 (배경색 없이)
    # 교통비 섹션 헤더
    for row in [8, 9]:
        for col in [1, 2, 3, 5]:  # A, B, C, E열
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and ('출장교통비' in cell.value or '이용차량' in cell.value or '대중교통' in cell.value):
                cell.font = header_font
    
    # 출장자 테이블 헤더 폰트 설정
    for col in range(1, 8):  # A~G열
        cell = ws.cell(row=11, column=col)
        if cell.value:
            cell.font = header_font
    
    # 행 높이 세밀 조정
    row_heights = {
        1: 60,   # 제목 - 살짝 늘림
        2: 30,   # 결재란
        3: 90,   # 결재란 - 더 늘림 (60→70)
        4: 40,   # 소속/과제책임자
        5: 40,   # 연구과제명
        6: 120,   # 출장기간/출장지
        7: 160,   # 출장목적 (긴 텍스트)
        8: 40,   # 교통비 헤더
        9: 40,   # 교통비 상세
        10: 40,  # 출장자 헤더
        11: 32,  # 출장자 테이블 헤더 (25+2)
        20: 40,  # 신청 문구 (25+5)
        21: 40,  # 날짜 (25+5)
        22: 60,  # 회사명 (25+5)
    }
    
    # 출장자 데이터 행들 (12-19행 - 8명까지) 2pt씩 늘림
    for row_num in range(12, 20):
        row_heights[row_num] = 32  # 30
    
    # 행 높이 적용
    for row_num, height in row_heights.items():
        ws.row_dimensions[row_num].height = height
    
    # 열 너비 최적화
    column_widths = {
        'A': 15,  # 직급/라벨
        'B': 15,  # 병합용
        'C': 15,  # 병합용
        'D': 15,  # 성명/값 (18→15로 줄임)
        'E': 19,  # 계좌번호/라벨 (동일 간격)
        'F': 19,  # 병합용 (동일 간격)
        'G': 19,  # 비고 (동일 간격)
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

def setup_application_page_settings(ws):
    """출장신청서용 페이지 설정"""
    
    # 페이지 방향 설정 (세로)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    
    # 용지 크기 설정 (A4)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # 여백 설정 (단위: 인치) - 더 적절한 여백
    ws.page_margins = PageMargins(
        left=0.7, right=0.7, top=0.8, bottom=0.8,
        header=0.3, footer=0.3
    )
    
    # 페이지에 맞춤 설정
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1  # 1페이지에 맞춤
    ws.page_setup.fitToWidth = 1   # 가로 1페이지
    
    # 인쇄 영역 명확하게 표시하기 위한 설정
    ws.page_setup.blackAndWhite = False  # 컬러 모드
    ws.page_setup.draft = False  # 초안 모드 비활성화
    
    # 인쇄영역 설정 (출장신청서는 A1:G22 - 8명까지 확장)
    ws.print_area = 'A1:G22'
    
    # 인쇄 범위 외부를 회색으로 표시하기 위한 설정
    ws.page_setup.scale = 100  # 배율 100%
    
    # 격자선 인쇄 설정
    ws.print_options.gridLines = False
    ws.print_options.gridLinesSet = True
    
    # 중앙 정렬
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    
    # 인쇄 제목 설정 (반복 인쇄할 행/열)
    ws.print_title_rows = '1:1'  # 1행 제목 반복
    
    # 페이지 나누기 미리보기 설정
    ws.sheet_view.view = 'pageBreakPreview'  # 페이지 나누기 미리보기
    ws.sheet_view.zoomScale = 100
    
    # 인쇄 품질 향상
    ws.page_setup.horizontalDpi = 300
    ws.page_setup.verticalDpi = 300
